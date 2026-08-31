#!/usr/bin/env python3
"""
Financial model generator.

One script, one model per business, driven by a spec dict. Written as a
generator rather than a hand-built workbook for the same reason the site is
generated: seven hand-made spreadsheets means seven places for a formula to rot.

Every cell that matters is a LIVE EXCEL FORMULA, not a value computed in Python
and pasted in. Change a dial on the Dials sheet and the whole workbook
recalculates in Excel, Numbers or Sheets. A model you cannot interrogate is a
slide, not a model.

Usage: python3 model/make_model.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# ---------------------------------------------------------------- styling

INK = "FF191A1C"
ACCENT = "FF1D4E6F"
WARN = "FFA83A2C"
H_FILL = PatternFill("solid", fgColor="FFE2EBF1")
DIAL_FILL = PatternFill("solid", fgColor="FFFFF4D6")   # editable cells are yellow
CALC_FILL = PatternFill("solid", fgColor="FFF7F7F5")
thin = Side(style="thin", color="FFD6D4CD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, cols, title=None):
    if title:
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, size=13, color=ACCENT)
        row += 1
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = Font(bold=True, size=9, color=INK)
        c.fill = H_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    return row + 1


def build(spec):
    wb = Workbook()

    # ============================================================ DIALS
    ws = wb.active
    ws.title = "Dials"
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 62

    t = ws.cell(row=1, column=1, value=f"{spec['name']} — model dials")
    t.font = Font(bold=True, size=16, color=ACCENT)
    ws.cell(row=2, column=1,
            value="Yellow cells are yours to change. Everything else is a formula. "
                  "Nothing on the other sheets is typed in by hand.").font = Font(italic=True, size=10)

    r = 4
    r = hdr(ws, r, ["Dial", "Value", "Note — where this number came from"])
    dial_rows = {}
    for key, label, value, note, fmt in spec["dials"]:
        ws.cell(row=r, column=1, value=label).border = BORDER
        c = ws.cell(row=r, column=2, value=value)
        c.fill = DIAL_FILL; c.border = BORDER; c.number_format = fmt
        c.font = Font(bold=True)
        n = ws.cell(row=r, column=3, value=note)
        n.border = BORDER; n.alignment = Alignment(wrap_text=True, vertical="top"); n.font = Font(size=9)
        dial_rows[key] = r
        r += 1

    D = lambda k: f"Dials!$B${dial_rows[k]}"

    # ============================================================ MONTHLY
    m = wb.create_sheet("Monthly")
    m.column_dimensions["A"].width = 9
    for i in range(2, 14):
        m.column_dimensions[get_column_letter(i)].width = 14

    cols = ["Month", "New adds", "Churned", "Customers", "MRR", "ARR run-rate",
            "COGS", "Gross profit", "Fixed costs", "Net", "Cumulative", "% of break-even"]
    row = hdr(m, 1, cols, f"{spec['name']} — 36 months")
    first = row

    for i in range(36):
        rr = row + i
        m.cell(row=rr, column=1, value=i + 1).border = BORDER
        prev_cust = f"D{rr-1}" if i else "0"

        # New adds ramp linearly from starting to steady rate over the ramp period,
        # then hold. Written as an Excel formula so the ramp is visible and editable.
        m.cell(row=rr, column=2,
               value=f"=ROUND(MIN({D('adds_steady')},{D('adds_start')}+"
                     f"({D('adds_steady')}-{D('adds_start')})*MIN(1,(A{rr}-1)/{D('ramp')})),0)")
        m.cell(row=rr, column=3, value=f"=ROUND({prev_cust}*{D('churn')},1)")
        m.cell(row=rr, column=4, value=f"=MAX(0,{prev_cust}+B{rr}-C{rr})")
        m.cell(row=rr, column=5, value=f"=D{rr}*{D('arpu')}")
        m.cell(row=rr, column=6, value=f"=E{rr}*12")
        m.cell(row=rr, column=7, value=f"=D{rr}*{D('cogs')}")
        m.cell(row=rr, column=8, value=f"=E{rr}-G{rr}")
        m.cell(row=rr, column=9, value=f"={D('fixed')}")
        m.cell(row=rr, column=10, value=f"=H{rr}-I{rr}")
        m.cell(row=rr, column=11, value=(f"=J{rr}" if i == 0 else f"=K{rr-1}+J{rr}"))
        m.cell(row=rr, column=12, value=f"=IF({D('target_mrr')}=0,0,E{rr}/{D('target_mrr')})")

        for cidx in range(1, 13):
            c = m.cell(row=rr, column=cidx)
            c.border = BORDER
            if cidx in (5, 6, 7, 8, 9, 10, 11):
                c.number_format = '"$"#,##0'
            if cidx == 12:
                c.number_format = "0%"
            if cidx in (2, 3, 4):
                c.number_format = "#,##0.0" if cidx == 3 else "#,##0"
            if cidx > 1:
                c.fill = CALC_FILL
    last = row + 35

    # chart — customers and MRR over time
    ch = LineChart()
    ch.title = "Customers and MRR"
    ch.height, ch.width = 8, 18
    data = Reference(m, min_col=4, max_col=5, min_row=first - 1, max_row=last)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(Reference(m, min_col=1, min_row=first, max_row=last))
    m.add_chart(ch, "N3")

    # ============================================================ SUMMARY
    s = wb.create_sheet("Summary", 1)
    s.column_dimensions["A"].width = 48
    s.column_dimensions["B"].width = 18
    s.column_dimensions["C"].width = 60

    s.cell(row=1, column=1, value=f"{spec['name']} — the questions that matter").font = Font(bold=True, size=16, color=ACCENT)
    r = 3
    r = hdr(s, r, ["Question", "Answer", "Why it matters"])
    for label, formula, note, fmt in [
        ("Customers needed to hit the target MRR",
         f"=CEILING({D('target_mrr')}/{D('arpu')},1)", "The single number the whole plan aims at.", "#,##0"),
        ("Months to reach the target",
         f'=IFERROR(MATCH(TRUE,INDEX(Monthly!$E${first}:$E${last}>={D("target_mrr")},0),0),"not within 36")',
         "If this says 'not within 36', the growth or churn dial is unrealistic.", "General"),
        ("Steady-state ceiling (adds ÷ churn)",
         f"=IFERROR({D('adds_steady')}/{D('churn')},0)",
         "Where customer count stops growing. IF THIS IS BELOW THE BREAK-EVEN COUNT, THE BUSINESS NEVER GETS THERE — churn eats it first. The most important cell in the workbook.", "#,##0"),
        ("Ceiling as a share of the target", f"=IFERROR({D('adds_steady')}/{D('churn')}/CEILING({D('target_mrr')}/{D('arpu')},1),0)",
         "Above 100% means the target is reachable at these dials. Below means it is not.", "0%"),
        ("MRR at month 12", f"=Monthly!E{first+11}", "", '"$"#,##0'),
        ("MRR at month 24", f"=Monthly!E{first+23}", "", '"$"#,##0'),
        ("MRR at month 36", f"=Monthly!E{last}", "", '"$"#,##0'),
        ("Cumulative cash at month 36", f"=Monthly!K{last}", "Negative means the business has not repaid its fixed costs.", '"$"#,##0'),
        ("Gross margin", f"=IFERROR(1-{D('cogs')}/{D('arpu')},0)",
         "Below about 70% and the flat-price model is under strain.", "0%"),
        ("Average customer lifetime (months)", f"=IFERROR(1/{D('churn')},0)",
         "The reciprocal of monthly churn.", "#,##0.0"),
        ("Lifetime value", f"=IFERROR(({D('arpu')}-{D('cogs')})/{D('churn')},0)",
         "Gross-margin LTV. Compare to what an acquisition is allowed to cost.", '"$"#,##0'),
        ("Market share needed at the target", f"=IFERROR(CEILING({D('target_mrr')}/{D('arpu')},1)/{D('market')},0)",
         spec["market_note"], "0.0%"),
    ]:
        s.cell(row=r, column=1, value=label).border = BORDER
        c = s.cell(row=r, column=2, value=formula)
        c.border = BORDER; c.number_format = fmt; c.font = Font(bold=True, color=ACCENT)
        n = s.cell(row=r, column=3, value=note)
        n.border = BORDER; n.font = Font(size=9); n.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # ============================================================ SENSITIVITY
    v = wb.create_sheet("Sensitivity")
    v.column_dimensions["A"].width = 22
    v.cell(row=1, column=1, value="Steady-state customer ceiling: adds per month ÷ monthly churn").font = Font(bold=True, size=13, color=ACCENT)
    v.cell(row=2, column=1, value="Churn across the top, monthly adds down the side. Red = below the break-even count, "
                                  "which means the business never reaches its target no matter how long it runs.").font = Font(italic=True, size=10)
    churns = [0.02, 0.03, 0.04, 0.05, 0.06, 0.08, 0.10]
    adds = [2, 3, 4, 5, 6, 8, 10, 12, 15]
    r0 = 4
    v.cell(row=r0, column=1, value="adds \\ churn").fill = H_FILL
    for j, ch_ in enumerate(churns):
        c = v.cell(row=r0, column=2 + j, value=ch_)
        c.number_format = "0%"; c.fill = H_FILL; c.font = Font(bold=True); c.border = BORDER
        v.column_dimensions[get_column_letter(2 + j)].width = 11
    for i, a in enumerate(adds):
        rr = r0 + 1 + i
        c = v.cell(row=rr, column=1, value=a); c.fill = H_FILL; c.font = Font(bold=True); c.border = BORDER
        for j in range(len(churns)):
            cc = v.cell(row=rr, column=2 + j,
                        value=f"=ROUND($A{rr}/{get_column_letter(2+j)}${r0},0)")
            cc.number_format = "#,##0"; cc.border = BORDER
    v.cell(row=r0 + len(adds) + 2, column=1,
           value=f"Break-even customer count for reference: {spec['breakeven_hint']}").font = Font(bold=True, color=WARN)

    # ============================================================ NOTES
    n = wb.create_sheet("Notes")
    n.column_dimensions["A"].width = 110
    n.cell(row=1, column=1, value="What is measured, what is assumed").font = Font(bold=True, size=14, color=ACCENT)
    rr = 3
    for line in spec["notes"]:
        c = n.cell(row=rr, column=1, value=line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("["):
            c.font = Font(bold=True, color=WARN)
        rr += 2

    return wb


# ---------------------------------------------------------------- specs

COUNTERCITE = {
    "name": "Countercite",
    "market_note": "Denominator is 1,708 active Texas public adjuster licences (TDI file, 2026-08-28). "
                   "Texas alone, one of three target professions, before Illinois or Florida.",
    "breakeven_hint": "101 customers at $299/mo = $30,199/mo",
    "dials": [
        ("arpu",       "Average revenue per customer per month", 299, "Practice tier. Blended will fall between Solo $149 and Firm $599.", '"$"#,##0'),
        ("cogs",       "Cost to serve one customer per month",    45, "COSTED at ~$18.50/customer (2026-08-31). Kept at 45 deliberately as headroom, because flat pricing absorbs heavy users.", '"$"#,##0'),
        ("fixed",      "Fixed costs per month",                  400, "Domain, hosting, tooling, accounting. Excludes your own time and the one-off legal spend.", '"$"#,##0'),
        ("adds_start", "New customers in month 1",                 3, "Founder outreach, by hand, from the TDI register.", "#,##0"),
        ("adds_steady","New customers per month at steady state",  9, "[ASSUMPTION] After SEO and referral compound. Unproven.", "#,##0"),
        ("ramp",       "Months to reach the steady rate",          9, "How long the channel takes to work.", "#,##0"),
        ("churn",      "Monthly churn",                        0.045, "[ASSUMPTION] 4.5%/mo = ~22 month average life. Solo professional software churns hard.", "0.0%"),
        ("target_mrr", "Target MRR",                           30000, "Midpoint of the $20-50k/month band.", '"$"#,##0'),
        ("market",     "Addressable customers (Texas PAs only)", 1708, "MEASURED, from the TDI licence file.", "#,##0"),
    ],
    "notes": [
        "[MEASURED] 1,708 active Texas public adjuster licences as of 2026-08-28, from the Texas Department of Insurance open data file. 981 resident in Texas, 727 licensed here but living elsewhere, 293 of those in Florida. Houston has 112 resident licensees.",
        "[MEASURED] New licences issued: 160 in 2023, 323 in 2024, 274 in 2025, 187 to August 2026. The 2024-25 cohort is 597 people and is the primary launch target.",
        "[RECOSTED 2026-08-31] Real cost to serve is about $10.60/customer/month at 100 customers on the $299 Practice tier: $1.20 of model inference, $8.97 of card processing, and about $0.47 of everything else. The inference figure is computed in api_cost.py from the pipeline the engine actually runs, not estimated -- retrieval is deterministic, so only candidate passage pairs reach a model and cost scales with pairs rather than page count. It replaces an earlier $9 guess that was 7.5x too high. Note what that leaves: card processing is now 85% of cost to serve, so annual billing and ACH on the Firm tier matter more to margin than any model choice. The dial stays at $45 as deliberate headroom, because flat pricing absorbs heavy users and a firm running thousand-page productions costs several times a light one. If the true figure were $120, gross margin would fall from 94% to 60% and the price would need to rise.",
        "[ASSUMPTION] Steady-state adds of 9/month. No channel has been run yet. This is the second most fragile number here.",
        "[ASSUMPTION] Churn of 4.5%/month. No data. Solo-professional software is generally worse than team software, and a public adjuster with no storm has no cases in March.",
        "[NOT ESTABLISHED] Special education advocates and patient advocates cannot be counted at all -- neither is a licensed profession with a public register. The market dial covers Texas public adjusters ONLY, which makes the market-share figure conservative but also means two of three segments are unmodelled.",
        "[CAUTION] Do not add state registries together. An adjuster licensed in Texas is frequently also licensed in Florida and Illinois. The national de-duplicated total is much smaller than the sum of the states.",
        "[THE CELL TO WATCH] Summary!B6, the steady-state ceiling. Adds divided by churn. If that number is below the break-even customer count, the business never reaches its target however long it runs -- churn eats it first. At 9 adds and 4.5% churn the ceiling is 200, comfortably above 101. At 5 adds and 6% churn it is 83, and the business fails. That is the whole risk in one cell.",
    ],
}

if __name__ == "__main__":
    import os
    out = os.path.join(os.path.dirname(__file__), "Countercite-Model.xlsx")
    build(COUNTERCITE).save(out)
    print("wrote", out)
