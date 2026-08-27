#!/usr/bin/env python3
"""Financial model for the five buildouts.

Dials drive everything. Change a price or a churn rate and every projection,
LTV and break-even figure moves with it.
"""
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = pathlib.Path(__file__).resolve().parent / "Five-Business-Models.xlsx"

ARIAL = "Arial"
BLUE = Font(name=ARIAL, size=10, color="0000FF")            # hardcoded input
BLACK = Font(name=ARIAL, size=10)                            # formula
GREEN = Font(name=ARIAL, size=10, color="008000")            # link to another sheet
BOLD = Font(name=ARIAL, size=10, bold=True)
H1 = Font(name=ARIAL, size=14, bold=True)
H2 = Font(name=ARIAL, size=11, bold=True, color="1F4E79")
HEADROW = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")

YELLOW = PatternFill("solid", fgColor="FFFF00")              # key assumption
DARK = PatternFill("solid", fgColor="16191D")
BAND = PatternFill("solid", fgColor="F2F5F8")

MONEY = '$#,##0;($#,##0);-'
MONEY2 = '$#,##0.00;($#,##0.00);-'
PCT1 = '0.0%'
PCT2 = '0.00%'
NUM = '#,##0;(#,##0);-'

thin = Side(style="thin", color="D9DEE5")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

# idea key, label, price, churn, market size (None = unknown), new/month
IDEAS = [
    ("Barrier",    "7 - Barrier survey",      79,  0.04, None,   4),
    ("Wall",       "8 - Behind the wall",     49,  0.04, 127394, 6),
    ("Templating", "1 - Templating",         119,  0.04, 5951,   3),
    ("FloorPrep",  "6 - Floor prep",          39,  0.04, 13108,  7),
    ("Fire",       "2 - Fire & life-safety",  89,  0.025, 19845, 3),
]
FIRST = 10                      # first idea row on Dials
MONTHS = 24


def head(ws, row, labels, widths=None):
    for i, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = HEADROW
        c.fill = DARK
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def build():
    wb = Workbook()

    # ---------------------------------------------------------------- Read Me
    ws = wb.active
    ws.title = "Read Me"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 92
    ws["A1"] = "Five business models - financial workbook"
    ws["A1"].font = H1
    rows = [
        ("", ""),
        ("What this is", "One model behind the five buildouts. Every figure on every other tab is a "
                         "formula driven by the Dials tab."),
        ("How to use it", "Edit ONLY the blue cells on 'Dials'. Everything else recalculates."),
        ("", ""),
        ("Colour legend", ""),
        ("  Blue text", "A hardcoded input or scenario lever. These are the cells you edit."),
        ("  Black text", "A formula. Do not overwrite."),
        ("  Green text", "A link pulling a value from another sheet."),
        ("  Yellow fill", "A key assumption that materially changes the answer."),
        ("", ""),
        ("Tabs", ""),
        ("  Dials", "Every input: commission, target, price, churn, market size, sales rate."),
        ("  Summary", "The five compared: net per customer, LTV, break-even, share of market."),
        ("  Barrier ... Fire", "A 24-month projection per idea."),
        ("", ""),
        ("Sourcing", "Apple's 15% commission under the Small Business Program (up to $1M proceeds) "
                     "is read off Apple's developer page [vendor]."),
        ("", "SMB SaaS monthly churn benchmarks of 3-5% are third-party reported and NOT "
             "independently confirmed [review]."),
        ("", "Market counts are third-party reported [review]; sources are listed in SOURCES.md."),
        ("", "Churn rates, sales rates and the $6,000 target are ASSUMPTIONS. Nobody has paid "
             "for any of these products, so no figure here is measured."),
        ("", ""),
        ("The one caveat", "Break-even is unaffected by churn; the replacement treadmill is. At 8% "
                           "monthly churn you replace twice the customers for the same revenue."),
    ]
    r = 2
    for a, b in rows:
        ws.cell(row=r, column=1, value=a).font = BOLD if a and not a.startswith("  ") else Font(name=ARIAL, size=10)
        c = ws.cell(row=r, column=2, value=b)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    for lbl_row, fnt in (("  Blue text", BLUE), ("  Black text", BLACK), ("  Green text", GREEN)):
        for rr in range(2, r):
            if ws.cell(row=rr, column=1).value == lbl_row:
                ws.cell(row=rr, column=1).font = fnt
    for rr in range(2, r):
        if ws.cell(row=rr, column=1).value == "  Yellow fill":
            ws.cell(row=rr, column=1).fill = YELLOW

    # ------------------------------------------------------------------ Dials
    d = wb.create_sheet("Dials")
    d["A1"] = "Dials - edit the blue cells only"
    d["A1"].font = H1
    d["A3"] = "Global"
    d["A3"].font = H2

    glob = [
        ("Apple commission", 0.15, PCT1,
         "15% under the App Store Small Business Program, up to $1M proceeds [vendor]. "
         "Rises to 30% above that threshold."),
        ("Monthly net target", 6000, MONEY,
         "What you need the business to pay you, net of Apple. ASSUMPTION."),
        ("Projection start customers", 0, NUM, "Everything starts at zero paying customers."),
    ]
    r = 4
    for label, val, fmt, note in glob:
        d.cell(row=r, column=1, value=label).font = BOLD
        c = d.cell(row=r, column=2, value=val)
        c.font = BLUE; c.number_format = fmt; c.fill = YELLOW; c.border = BOX
        n = d.cell(row=r, column=3, value=note)
        n.font = Font(name=ARIAL, size=9, italic=True, color="5B6470")
        n.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    d["A8"] = "Per idea"
    d["A8"].font = H2
    head(d, 9,
         ["Idea", "Price / month", "Monthly churn", "US market size",
          "New customers / month", "Notes"],
         [26, 14, 14, 16, 18, 60])

    for i, (key, label, price, churn, market, newpm) in enumerate(IDEAS):
        row = FIRST + i
        d.cell(row=row, column=1, value=label).font = BOLD
        for col, val, fmt in ((2, price, MONEY), (3, churn, PCT1),
                              (4, market, NUM), (5, newpm, NUM)):
            c = d.cell(row=row, column=col)
            if val is not None:
                c.value = val
            c.font = BLUE; c.number_format = fmt; c.border = BOX
            if col in (2, 3):
                c.fill = YELLOW
        note = {
            "Barrier": "Market size UNKNOWN - the CASp register was not reachable and no national "
                       "count was found. Counting it is task zero.",
            "Wall": "127,394 residential remodelers, 2020 Census [review]. Largest denominator.",
            "Templating": "5,951 countertop businesses [review]. Price raised from $59 to $119 "
                          "because 2% market share was needed at $59.",
            "FloorPrep": "13,108 flooring contractors [review]. Churn likely worse than 4% - "
                         "lowest usage frequency of the five.",
            "Fire": "19,845 fire protection & security contractors [review], likely ~half are "
                    "fire-only. Churn set to 2.5%: the frequency tracker holds customers.",
        }[key]
        c = d.cell(row=row, column=6, value=note)
        c.font = Font(name=ARIAL, size=9, italic=True, color="5B6470")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        d.row_dimensions[row].height = 30

    d.cell(row=FIRST + len(IDEAS) + 1, column=1,
           value="Blue = edit. Yellow = the assumptions that most change the answer.").font = \
        Font(name=ARIAL, size=9, italic=True)

    # ---------------------------------------------------------------- Summary
    s = wb.create_sheet("Summary")
    s["A1"] = "Summary - all five compared"
    s["A1"].font = H1
    s["A2"] = "Every cell below is a formula driven by Dials. Nothing here is typed."
    s["A2"].font = Font(name=ARIAL, size=9, italic=True, color="5B6470")

    head(s, 4,
         ["Idea", "Price / mo", "Net / mo\nafter Apple", "Monthly\nchurn",
          "Lifetime\n(months)", "LTV", "Break-even\ncustomers",
          "US market\nsize", "Share of market\nneeded",
          "Replacements\n/ month"],
         [26, 12, 13, 11, 12, 13, 13, 13, 16, 14])

    for i in range(len(IDEAS)):
        row = 5 + i
        dr = FIRST + i
        s.cell(row=row, column=1, value=f"=Dials!A{dr}").font = GREEN
        s.cell(row=row, column=2, value=f"=Dials!B{dr}").font = GREEN
        s.cell(row=row, column=2).number_format = MONEY
        s.cell(row=row, column=3, value=f"=Dials!B{dr}*(1-Dials!$B$4)").font = BLACK
        s.cell(row=row, column=3).number_format = MONEY2
        s.cell(row=row, column=4, value=f"=Dials!C{dr}").font = GREEN
        s.cell(row=row, column=4).number_format = PCT1
        s.cell(row=row, column=5, value=f"=IF(Dials!C{dr}=0,0,1/Dials!C{dr})").font = BLACK
        s.cell(row=row, column=5).number_format = '#,##0.0'
        s.cell(row=row, column=6, value=f"=IF(Dials!C{dr}=0,0,C{row}/Dials!C{dr})").font = BLACK
        s.cell(row=row, column=6).number_format = MONEY
        s.cell(row=row, column=7, value=f"=IF(C{row}=0,0,ROUNDUP(Dials!$B$5/C{row},0))").font = BLACK
        s.cell(row=row, column=7).number_format = NUM
        s.cell(row=row, column=8, value=f"=Dials!D{dr}").font = GREEN
        s.cell(row=row, column=8).number_format = NUM
        s.cell(row=row, column=9,
               value=f'=IF(ISNUMBER(Dials!D{dr}),IF(Dials!D{dr}=0,"n/a",G{row}/Dials!D{dr}),"unknown")').font = BLACK
        s.cell(row=row, column=9).number_format = PCT2
        s.cell(row=row, column=10, value=f"=ROUNDUP(G{row}*Dials!C{dr},0)").font = BLACK
        s.cell(row=row, column=10).number_format = NUM
        for col in range(1, 11):
            s.cell(row=row, column=col).border = BOX
            if i % 2 == 1:
                s.cell(row=row, column=col).fill = BAND

    n = 5 + len(IDEAS) + 1
    s.cell(row=n, column=1,
           value="Break-even is the customer count that hits the monthly net target. It does NOT "
                 "move with churn - the replacement rate in the last column does.").font = \
        Font(name=ARIAL, size=9, italic=True, color="5B6470")
    s.cell(row=n + 1, column=1,
           value="'Share of market needed' is the honest ranking. Idea 1 needs ~18x the "
                 "penetration idea 8 does.").font = \
        Font(name=ARIAL, size=9, italic=True, color="5B6470")
    s.cell(row=n + 2, column=1,
           value="Barrier survey shows 'unknown' because its denominator has not been "
                 "established. That is the risk, not a missing cell.").font = \
        Font(name=ARIAL, size=9, italic=True, color="A32D1E")

    # ------------------------------------------------------- projection tabs
    for i, (key, label, *_rest) in enumerate(IDEAS):
        dr = FIRST + i
        p = wb.create_sheet(key)
        p["A1"] = f"{label} - 24-month projection"
        p["A1"].font = H1
        p["A2"] = ("Driven entirely by Dials. Churned customers are rounded to whole people, "
                   "so the curve steps rather than glides.")
        p["A2"].font = Font(name=ARIAL, size=9, italic=True, color="5B6470")

        p["A4"] = "Price / mo";      p["B4"] = f"=Dials!B{dr}";  p["B4"].font = GREEN; p["B4"].number_format = MONEY
        p["A5"] = "Monthly churn";   p["B5"] = f"=Dials!C{dr}";  p["B5"].font = GREEN; p["B5"].number_format = PCT1
        p["A6"] = "New / month";     p["B6"] = f"=Dials!E{dr}";  p["B6"].font = GREEN; p["B6"].number_format = NUM
        p["A7"] = "Apple commission"; p["B7"] = "=Dials!$B$4";   p["B7"].font = GREEN; p["B7"].number_format = PCT1
        for rr in range(4, 8):
            p.cell(row=rr, column=1).font = BOLD
            p.cell(row=rr, column=2).border = BOX

        head(p, 9,
             ["Month", "Customers\nstart", "New", "Churned", "Customers\nend",
              "Gross MRR", "Net MRR\nafter Apple", "Cumulative\nnet"],
             [9, 12, 9, 11, 12, 14, 14, 14])

        for m in range(1, MONTHS + 1):
            row = 9 + m
            p.cell(row=row, column=1, value=m).font = BLACK
            if m == 1:
                p.cell(row=row, column=2, value="=Dials!$B$6").font = GREEN
            else:
                p.cell(row=row, column=2, value=f"=E{row-1}").font = BLACK
            p.cell(row=row, column=3, value="=$B$6").font = BLACK
            p.cell(row=row, column=4, value=f"=ROUND(B{row}*$B$5,0)").font = BLACK
            p.cell(row=row, column=5, value=f"=B{row}+C{row}-D{row}").font = BLACK
            p.cell(row=row, column=6, value=f"=E{row}*$B$4").font = BLACK
            p.cell(row=row, column=7, value=f"=F{row}*(1-$B$7)").font = BLACK
            if m == 1:
                p.cell(row=row, column=8, value=f"=G{row}").font = BLACK
            else:
                p.cell(row=row, column=8, value=f"=H{row-1}+G{row}").font = BLACK
            for col, fmt in ((2, NUM), (3, NUM), (4, NUM), (5, NUM),
                             (6, MONEY), (7, MONEY), (8, MONEY)):
                p.cell(row=row, column=col).number_format = fmt
            for col in range(1, 9):
                p.cell(row=row, column=col).border = BOX
                if m % 2 == 0:
                    p.cell(row=row, column=col).fill = BAND

        last = 9 + MONTHS
        p.cell(row=last + 2, column=1, value="Month 24 net MRR").font = BOLD
        p.cell(row=last + 2, column=2, value=f"=G{last}").font = BLACK
        p.cell(row=last + 2, column=2).number_format = MONEY
        p.cell(row=last + 3, column=1, value="Break-even reached in month").font = BOLD
        p.cell(row=last + 3, column=2,
               value=f'=IFERROR(MATCH(TRUE,INDEX(G10:G{last}>=Dials!$B$5,0),0),"not within 24 months")').font = BLACK
        p.cell(row=last + 4, column=1, value="Steady-state customers at this sales rate").font = BOLD
        p.cell(row=last + 4, column=2, value="=IF($B$5=0,0,$B$6/$B$5)").font = BLACK
        p.cell(row=last + 4, column=2).number_format = NUM
        p.cell(row=last + 5, column=1,
               value="Steady state = new per month / churn rate. Beyond it, this sales rate "
                     "cannot grow the business further.").font = \
            Font(name=ARIAL, size=9, italic=True, color="5B6470")
        for rr in (last + 2, last + 3, last + 4):
            p.cell(row=rr, column=2).border = BOX

    wb.save(OUT)
    print(f"wrote {OUT}  ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    build()
